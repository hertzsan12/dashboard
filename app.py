import streamlit as st
import pandas as pd
import datetime
import os
import hashlib
from openpyxl import Workbook, load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ---------- File Paths ----------
BASE_DIR = "."
TRANSACTIONS_FILE = os.path.join(BASE_DIR, 'transactions_log.xlsx')
EQUIPMENT_FILE = os.path.join(BASE_DIR, 'equipment_stock.xlsx')
AUDIT_FILE = os.path.join(BASE_DIR, 'registration_audit.xlsx')
CREDENTIALS_FILE = os.path.join(BASE_DIR, 'user_credentials.xlsx')

LOW_STOCK_THRESHOLD = 5
MAX_STOCK_THRESHOLD = 20  # Example max stock threshold (can be adjusted or read from file)

# ---------- Authentication ----------
def connect_gsheet():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]

    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        st.secrets["gcp_credentials"], scope
    )

    client = gspread.authorize(creds)
    return client
    
def ensure_workbook(file, headers):
    return  # Disable file creation in cloud

def load_user_credentials():
    ensure_workbook(CREDENTIALS_FILE, ['Username', 'Password', 'Role'])
    wb = load_workbook(CREDENTIALS_FILE)
    ws = wb.active
    credentials = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        username, password_hash, role = row
        if username and password_hash:
            credentials[username.upper()] = {"password": password_hash, "role": role}
    return credentials

def save_user_credentials(username, password, role="viewer"):
    ensure_workbook(CREDENTIALS_FILE, ['Username', 'Password', 'Role'])
    wb = load_workbook(CREDENTIALS_FILE)
    ws = wb.active
    hashed_pw = hashlib.sha256(password.encode()).hexdigest()
    ws.append([username.upper(), hashed_pw, role])
    wb.save(CREDENTIALS_FILE)

def register_user(username, password):
    username = username.upper()
    creds = load_user_credentials()
    if username in creds:
        return False
    save_user_credentials(username, password)
    log_registration(username)
    return True

def authenticate(username, password):
    creds = load_user_credentials()
    user = creds.get(username.upper())
    if user:
        return hashlib.sha256(password.encode()).hexdigest() == user["password"]
    return False

def get_user_role(username):
    return "admin"

# ---------- Excel Utilities ----------
def read_inventory():

    client = connect_gsheet()
    
    sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("equipment_stock")

    st.success("Connected to Google Sheets!")

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    inventory = {}
    uoms = {}

    for _, row in df.iterrows():
        item = row.get("Item")
        qty = int(row.get("Qty", 0))
        uom = row.get("UOM", "pcs")

        if not item:
            continue

        inventory[item] = inventory.get(item, 0) + qty
        uoms[item] = uom

    return inventory, uoms
    
# =========================
# SAFE WORKBOOK LOADER
# =========================
def safe_load_workbook(file):
    try:
        return load_workbook(file)
    except Exception as e:
        st.error(f"Error loading {file}: {e}")
        return None


# =========================
# READ EQUIPMENT ITEMS
# =========================
def read_equipment_items():
    client = connect_gsheet()
    sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("equipment_stock")

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    equipment_dict = {}

    for _, row in df.iterrows():
        eq = row.get("Equipment")
        item = normalize_item_name(row.get("Item"))
        qty = int(row.get("Qty", 0))
        uom = row.get("UOM", "pcs")

        if not eq:
            continue

        if eq not in equipment_dict:
            equipment_dict[eq] = {}

        if item:
            if item not in equipment_dict[eq]:
                equipment_dict[eq][item] = {"qty": 0, "uom": uom}

            equipment_dict[eq][item]["qty"] += qty
            equipment_dict[eq][item]["uom"] = uom
        if item == "__RESET__":
            equipment_dict[eq] = {}
            continue

    return equipment_dict

# =========================
# WRITE EQUIPMENT ITEMS
# =========================
def write_equipment_items(equipment_dict):
    pass
# =========================
# APPEND EQUIPMENT STOCK (GOOGLE SHEETS)
# =========================
def append_equipment_stock(equipment, item, qty, uom="pcs"):
    client = connect_gsheet()

    sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("equipment_stock")

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet.append_row([
        timestamp,
        equipment,
        item,
        qty,
        uom
    ])

def normalize_item_name(name):
    if not name:
        return ""
    return name.upper().replace(" ", "")


# =========================
# LOG TRANSACTION (DISABLED)
# =========================
def log_transaction(action, item, quantity, person, mdr_number=None, equipment=None, uom="pcs"):
    client = connect_gsheet()

    sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("transactions_log")

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    qty_to_log = -quantity if action == "withdraw" else quantity

    sheet.append_row([
        timestamp,
        action,
        item,
        qty_to_log,
        uom,
        person,
        mdr_number if action == "deliver" else "",
        equipment
    ])

    st.success("✅ Transaction saved to Google Sheets!")

# ---------- Google Sheets Functions ----------

def update_equipment_stock(equipment, item, qty, uom):
    client = connect_gsheet()

    sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("equipment_stock")

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet.append_row([
        timestamp,
        equipment,
        item,
        qty,
        uom
    ])
    
# ---------- Streamlit App ----------
def force_rerun():
    st.session_state['rerun_counter'] = st.session_state.get('rerun_counter', 0) + 1

st.set_page_config(page_title="Plant Inventory Monitoring", layout="wide")

st.session_state.authenticated = True
st.session_state.username = "admin"

if not st.session_state.authenticated:
    st.title("Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if authenticate(username, password):
            st.session_state.authenticated = True
            st.session_state.username = username
            st.success("Login successful")
            force_rerun()
        else:
            st.error("Invalid credentials")

    st.write("Don't have an account?")
    new_user = st.text_input("New Username")
    new_pass = st.text_input("New Password", type="password")
    if st.button("Register"):
        if register_user(new_user, new_pass):
            st.success("User registered successfully")
        else:
            st.warning("Username already exists")

else:
    username = st.session_state.username
    user_role = get_user_role(username)
    is_admin = (user_role == "admin")

    st.sidebar.title("David Hertz Monitoring")
    choice = st.sidebar.radio("Go to", ["Inventory", "Equipment", "Withdraw/Deliver", "Transactions", "Logout"])

    if choice == "Logout":
        st.session_state.authenticated = False
        st.session_state.username = ""
        force_rerun()

    elif choice == "Inventory":
        st.title("Inventory Overview")
        inventory, uoms = read_inventory()
        data = []
        for item in inventory:
            qty = inventory[item]
            uom = uoms.get(item, "pcs")

            # Stock status logic:
            if qty == 0:
                status = "🔴 No Stock: Purchase?"
            elif qty <= LOW_STOCK_THRESHOLD:
                status = "🟡 Running out of Stock"
            elif qty <= MAX_STOCK_THRESHOLD:
                status = "🟢 Stock OK"
            else:
                status = "🟢 Stock OK"

            data.append({"Item": item, "Quantity": qty, "UOM": uom, "Status": status})

        df = pd.DataFrame(data, columns=["Item", "Quantity", "UOM", "Status"])
        st.dataframe(df, use_container_width=True)

    elif choice == "Equipment":
        st.title("Equipment Inventory")

        equipment_items = read_equipment_items()
        equipment_list = sorted(equipment_items.keys())
        options = ["-- New Equipment --"] + equipment_list

        selected_eq = st.selectbox("Select Equipment", options)
        is_new = selected_eq == "-- New Equipment --"

        eq_name = ""
        if is_new:
            eq_name = st.text_input("Enter new equipment name")
        else:
            eq_name = st.text_input("Edit equipment name", value=selected_eq)

        # ✅ SHOW ITEMS
        
            items = equipment_items.get(eq_name, {})

            df_items = pd.DataFrame(
                [{"Item": item, "Quantity": data["qty"], "UOM": data["uom"]}
                 for item, data in items.items()]
            )

            df_items = pd.concat([
                df_items,
                pd.DataFrame([{"Item": "", "Quantity": 0, "UOM": "pcs"}])
            ], ignore_index=True)

            # ✅ SAVE ITEMS (FULL FIX)
            if st.button("Save Equipment Items", key=f"save_items_{eq_name}"):
                if not is_admin:
                    st.warning("Only admins can edit equipment items.")
                else:
                    edited_df = edited_df.dropna(subset=['Item'])
                    edited_df = edited_df[edited_df['Item'] != ""]

                    updated_items = {}

                    for _, row in edited_df.iterrows():
                        item = normalize_item_name(row['Item'])
                        qty = int(row['Quantity']) if pd.notna(row['Quantity']) else 0
                        uom = row['UOM'] if pd.notna(row['UOM']) else "pcs"

                        updated_items[item] = {"qty": qty, "uom": uom}

                    # 🔥 RESET old values
                    append_equipment_stock(eq_name, "__RESET__", 0, "")

                    # 🔥 SAVE new values
                    for item, data in updated_items.items():
                        append_equipment_stock(eq_name, item, data["qty"], data["uom"])
    
                    st.success("Equipment items saved successfully.")
                    st.rerun()

        
            items = equipment_items.get(eq_name, {})
            df_items = pd.DataFrame(
                [{"Item": item, "Quantity": data["qty"], "UOM": data["uom"]} for item, data in items.items()]
            )

            df_items = pd.concat([df_items, pd.DataFrame([{"Item": "", "Quantity": 0, "UOM": "pcs"}])], ignore_index=True)

            st.markdown("### Edit Item Quantities and UOM")
            
            key_name = f"equip_edit_{eq_name}" if eq_name else "equip_edit_default"
            
            edited_df = st.data_editor(
                df_items,
                num_rows="dynamic",
                use_container_width=True,
                key=f"equip_edit_{eq_name}"
            )

            if st.button("Save Equipment Items"):
                if not is_admin:
                    st.warning("Only admins can edit equipment items.")
                else:
                    # Remove empty rows
                    edited_df = edited_df.dropna(subset=['Item'])
                    edited_df = edited_df[edited_df['Item'] != ""]
                    # Build new equipment dict
                    equipment_items[eq_name] = {}
                    # Build new equipment dict
                    updated_items = {}
                    for _, row in edited_df.iterrows():
                        item = normalize_item_name(row['Item'])
                        qty = int(row['Quantity']) if pd.notna(row['Quantity']) else 0
                        uom = row['UOM'] if pd.notna(row['UOM']) else "pcs"
                        updated_items[item] = {"qty": qty, "uom": uom}

                    # Update the main dictionary and write back to Excel
                    for item, data in updated_items.items():
                        append_equipment_stock(eq_name, item, data["qty"], data["uom"])

                    st.success("Equipment items saved to Google Sheets.")
                    st.rerun()

    elif choice == "Withdraw/Deliver":
        st.title("Withdraw or Deliver Items")
        if not is_admin:
            st.warning("Only admins can withdraw or deliver items.")
        else:
            equipment_items = read_equipment_items()
            equipment_list = sorted(equipment_items.keys())
            equipment_selected = st.selectbox("Select Equipment", equipment_list)

            if equipment_selected:
                items = equipment_items.get(equipment_selected, {})
                if not items:
                    st.info("No items available for this equipment.")
                else:
                    item_list = list(items.keys())
                    item_selected = st.selectbox("Select Item", item_list)

                    inventory, uoms = read_inventory()
                    total_qty = inventory.get(item_selected, 0)
                    uom = uoms.get(item_selected, "pcs")

                    st.write(f"Current Stock (Total): {total_qty} {uom}")

                    current_qty = items[item_selected]['qty']
                    st.write(f"Current Stock in '{equipment_selected}': {current_qty} {uom}")

                    if current_qty == 0:
                        if total_qty > 0:
                            st.warning("Withdraw Stocks from other Equipment")
                        else:
                            st.error("Follow up Purchase / MR")

                    action = st.radio("Action", ["Withdraw", "Deliver"])

                    if action == "Withdraw":
                        st.write(f"Maximum Withdrawal = {current_qty} {uom}")
                        qty = st.number_input("Quantity", min_value=0, max_value=current_qty, step=1)
                    else:
                        qty = st.number_input("Quantity", min_value=0, step=1)

                    person = st.text_input("Person in Charge")
                    mdr_number = None
                    if action == "Deliver":
                        mdr_number = st.text_input("MDR Number")
                    elif action == "Deliver" and not mdr_number.strip():
                        st.warning("Please enter MDR Number.")

                    # Disable submit button logic
                    disable_submit = (
                        (action == "Withdraw" and current_qty == 0) or
                        (action == "Withdraw" and qty == 0) or
                        (action == "Deliver" and qty == 0)
)

                    if disable_submit:
                        if current_qty == 0 and total_qty == 0:
                            st.error("Follow up Purchase / MR")
                        elif current_qty == 0 and total_qty > 0:
                            st.warning("Withdraw Stocks from other Equipment")

                    if st.button("Submit Transaction"):
                        if not person.strip():
                            st.warning("Please enter the person in charge.")
                        elif action == "Withdraw" and qty > current_qty:
                            st.warning(f"Cannot withdraw more than available quantity ({current_qty}).")
                        else:
                            if action == "Withdraw":
                                items[item_selected]['qty'] -= qty
                            else:
                                items[item_selected]['qty'] += qty

                            if action == "Withdraw":
                                update_equipment_stock(equipment_selected, item_selected, -qty, uom)
                            else:
                                update_equipment_stock(equipment_selected, item_selected, qty, uom)

                            # ✅ Log transaction
                            log_transaction(
                                action=action.lower(),
                                item=item_selected,
                                quantity=qty,
                                person=person,
                                mdr_number=mdr_number,
                                equipment=equipment_selected,
                                uom=uom
                            )

                            st.success(f"{action} successful.")

                            # 🔥 THIS IS THE KEY
                            st.rerun()

    elif choice == "Transactions":
        st.title("Transaction Log")

        client = connect_gsheet()
        sheet = client.open_by_key("1Z-DPnZlZqZsAGWdAT8S-a2RUN9tqR0rnOMs3519VbBg").worksheet("transactions_log")

        # 🔥 ALWAYS get latest data
        data = sheet.get_all_values()

        if len(data) > 1:
            headers = data[0]
            rows = data[1:]

            df_log = pd.DataFrame(rows, columns=headers)

            df_log.columns = df_log.columns.str.strip()

            if "Timestamp" in df_log.columns:
                df_log['Timestamp'] = pd.to_datetime(df_log['Timestamp'])
                df_log = df_log.sort_values(by='Timestamp', ascending=False).reset_index(drop=True)

            st.dataframe(df_log, use_container_width=True)

        else:
            st.info("No transactions logged yet.")
